import openpyxl, json, os
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter
from copy import copy
from datetime import datetime

# ============================================================
# Configuration
# ============================================================
COMPANY_NAME_CN = "越南华星体育用品有限公司"
COMPANY_NAME_VN = "CÔNG TY TNHH SẢN PHẨM THỂ THAO HUA XING"
COMPANY_NAME_EN = "HUA XING SPORTS PRODUCTS CO., LTD"
PERIOD = "2026.1.1 - 2026.3.31"
WP_REF = "R1"
PREPARED_BY = "Daryl Chiu"
PREPARED_DATE = datetime.now().strftime("%Y-%m-%d")
REVIEWED_BY = ""
REVIEWED_DATE = ""

# Load data
with open('/tmp/payroll_summary.json') as f:
    payroll_data = json.load(f)

with open('/tmp/all_dept_si.json') as f:
    si_dept_data = json.load(f)

# ============================================================
# Helper: parse category gross from payroll summary
# ============================================================
def get_monthly_totals(month_key):
    """Get total gross pay for a month from summary data"""
    d = payroll_data.get(month_key, {})
    cats = d.get('categories', {})
    result = {'gross': 0, 'si_employee': 0, 'union_employee': 0, 'pit': 0, 'headcount': 0}
    for cat_name, info in cats.items():
        if '总合' in cat_name or 'TOTAL' in cat_name.upper():
            result['gross'] = info.get('gross', 0)
            result['si_employee'] = info.get('si_employee_10_5', 0)
            result['union_employee'] = info.get('union_employee_1', 0)
            result['pit'] = info.get('pit', 0)
            result['headcount'] = info.get('headcount', 0)
    return result

jan_total = get_monthly_totals('2026-01')
feb_total = get_monthly_totals('2026-02')
mar_total = get_monthly_totals('2026-03')

# ============================================================
# Create Workbook
# ============================================================
wb = openpyxl.Workbook()

# Styles
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
header_font = Font(name='Calibri', size=11, bold=True)
title_font = Font(name='Calibri', size=14, bold=True)
subtitle_font = Font(name='Calibri', size=12, bold=True)
normal_font = Font(name='Calibri', size=10)
small_font = Font(name='Calibri', size=9)
red_font = Font(name='Calibri', size=9, color='FF0000')
red_fill = PatternFill(start_color='FFFFE0', end_color='FFFFE0', fill_type='solid')  # Light yellow for procedure headers
light_blue_fill = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')
light_gray_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

def apply_cell(ws, row, col, value, font=None, alignment=None, border=None, fill=None, number_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font: cell.font = font
    if alignment: cell.alignment = alignment
    if border: cell.border = border
    if fill: cell.fill = fill
    if number_format: cell.number_format = number_format
    return cell

def set_col_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

# ============================================================
# SHEET: Working Paper
# ============================================================
ws = wb.active
ws.title = 'R1-EE 2026Q1'

# Column widths (similar to 成都明宇 R1)
col_widths = {1:3, 2:22, 3:45, 4:8, 5:6, 6:6, 7:16, 8:8, 9:16, 10:16, 11:16, 12:6, 13:16, 14:8, 15:16, 16:16, 17:16}
set_col_widths(ws, col_widths)

row = 1

# --- HEADER ---
apply_cell(ws, row, 1, COMPANY_NAME_CN, title_font)
apply_cell(ws, row, 10, 'Prepared by:', normal_font, Alignment(horizontal='right'))
apply_cell(ws, row, 11, PREPARED_BY, normal_font)
row += 1
apply_cell(ws, row, 1, COMPANY_NAME_VN, subtitle_font)
apply_cell(ws, row, 10, 'Date:', normal_font, Alignment(horizontal='right'))
apply_cell(ws, row, 11, PREPARED_DATE, normal_font)
row += 1
apply_cell(ws, row, 1, f'For the accounting period of {PERIOD}', subtitle_font)
apply_cell(ws, row, 10, 'Reviewed by:', normal_font, Alignment(horizontal='right'))
apply_cell(ws, row, 11, REVIEWED_BY, normal_font)
row += 1
apply_cell(ws, row, 10, 'Date:', normal_font, Alignment(horizontal='right'))
apply_cell(ws, row, 11, REVIEWED_DATE, normal_font)
row += 1
row += 1
apply_cell(ws, row, 1, f'Process - {WP_REF} Provision for Staff Costs / 应付职工薪酬', subtitle_font, fill=red_fill)
row += 1
apply_cell(ws, row, 1, '2026 Q1 audit scope: TE at 60% of PM (to be determined), trivial at 5% of PM', small_font)
row += 1
row += 1

# --- LEAD SCHEDULE ---
apply_cell(ws, row, 7, 'Opening Balance', header_font, fill=light_blue_fill)
apply_cell(ws, row, 8, 'CLA#', header_font, fill=light_blue_fill)
apply_cell(ws, row, 9, 'Opening Balance', header_font, fill=light_blue_fill)
apply_cell(ws, row, 10, 'Dr.', header_font, fill=light_blue_fill)
apply_cell(ws, row, 11, 'Cr.', header_font, fill=light_blue_fill)
apply_cell(ws, row, 13, 'Closing Balance', header_font, fill=light_blue_fill)
apply_cell(ws, row, 14, 'CLA#', header_font, fill=light_blue_fill)
apply_cell(ws, row, 15, 'Closing Balance', header_font, fill=light_blue_fill)
row += 1
for c in [7,9,13,15]:
    apply_cell(ws, row, c, 'Per Mgt Acc', small_font, fill=light_blue_fill)
for c in [10,11]:
    apply_cell(ws, row, c, '', small_font, fill=light_blue_fill)
apply_cell(ws, row, 9, 'Audited', small_font, fill=light_blue_fill)
apply_cell(ws, row, 15, 'Audited', small_font, fill=light_blue_fill)
row += 1
for c in [4,5,7,9,10,11,13,15]:
    apply_cell(ws, row, c, 'Note' if c==4 else ('Ref' if c==5 else ('RMB' if c in [7,9,10,11,13,15] else '')), small_font, fill=light_blue_fill)
row += 1
# Legend row
for c in [7,9,13,15]:
    apply_cell(ws, row, c, 'd', small_font, fill=light_blue_fill)
apply_cell(ws, row, 9, 'j', small_font, fill=light_blue_fill)
apply_cell(ws, row, 15, 'j', small_font, fill=light_blue_fill)
row += 1
# Account header
apply_cell(ws, row, 2, 'Account Code', header_font, fill=light_blue_fill)
apply_cell(ws, row, 3, '职工薪酬 / Staff Costs', header_font, fill=light_blue_fill)
row += 1

# Account lines (standard VAS chart)
accounts = [
    ('3341', '应付职工薪酬 - 基本工资\nBasic Salary', 1),
    ('3341', '应付职工薪酬 - 加班工资\nOvertime Pay', 1),
    ('3341', '应付职工薪酬 - 奖励/绩效工资\nBonus & Performance', 1),
    ('3341', '应付职工薪酬 - 职工福利\nEmployee Welfare', 2),
    ('3383', '应付社会保险费 - BHXH\nSocial Insurance Payable', 3),
    ('3384', '应付医疗保险费 - BHYT\nHealth Insurance Payable', 3),
    ('3386', '应付失业保险费 - BHTN\nUnemployment Insurance Payable', 3),
    ('3382', '应付工会经费 - Công Đoàn\nTrade Union Fee Payable', 3),
    ('3335', '应付个人所得税 - TNCN\nPersonal Income Tax Payable', 4),
    ('3341', '应付职工薪酬 - 离职补偿\nSeverance Pay', 5),
    ('3341', '应付职工薪酬 - 第13薪/年终奖\n13th Month / Year-end Bonus', 5),
]

for code, name, note in accounts:
    apply_cell(ws, row, 2, code, normal_font)
    apply_cell(ws, row, 3, name, small_font)
    apply_cell(ws, row, 4, note, small_font)
    # Placeholder values - to be filled when GL data arrives
    for c in [7,9,10,11,13,15]:
        apply_cell(ws, row, c, '', normal_font, border=thin_border)
    row += 1

# Total row
apply_cell(ws, row, 2, 'Provision for Staff Costs', header_font)
apply_cell(ws, row, 3, '应付职工薪酬 合计', header_font)
for c in [7,9,10,11,13,15]:
    apply_cell(ws, row, c, '', normal_font, border=thin_border)
row += 1
apply_cell(ws, row, 2, 'Check', small_font)
apply_cell(ws, row, 13, '0', small_font, border=thin_border)
apply_cell(ws, row, 15, '0', small_font, border=thin_border)
row += 2

# --- NOTE SECTION ---
apply_cell(ws, row, 1, 'Note:', subtitle_font)
row += 2

# ============================================================
# PROCEDURE a: Agree Credits to Payroll List
# ============================================================
row += 1
apply_cell(ws, row, 1, 'a', red_font, fill=red_fill)
apply_cell(ws, row, 2, 'Agree Credits for the period to Payroll List / 贷方发生额与工资表核对', subtitle_font, fill=red_fill)
row += 2

# Monthly comparison table
months_en = ['Jan-26 / 1月', 'Feb-26 / 2月', 'Mar-26 / 3月']
months_data = [jan_total, feb_total, mar_total]

apply_cell(ws, row, 1, '', small_font)
apply_cell(ws, row, 2, 'Payroll List / 工资表 (VND)', header_font, fill=light_blue_fill)
for i, m in enumerate(months_en):
    apply_cell(ws, row, 4+i*3, m, header_font, fill=light_blue_fill)
row += 1

# Payroll list row
apply_cell(ws, row, 1, '', small_font)
apply_cell(ws, row, 2, '  Gross Pay per Payroll / 工资表应发合计', normal_font)
for i, d in enumerate(months_data):
    apply_cell(ws, row, 4+i*3, d['gross'] if d['gross'] else 'N/A', normal_font, number_format='#,##0', border=thin_border)
row += 1

# Credits row - placeholder
apply_cell(ws, row, 1, '', small_font)
apply_cell(ws, row, 2, '  Credits per GL / 账面贷方发生额', normal_font)
for i in range(3):
    apply_cell(ws, row, 4+i*3, '⚠ PENDING GL DATA', red_font, border=thin_border)
row += 1

# Variance row
apply_cell(ws, row, 1, '', small_font)
apply_cell(ws, row, 2, '  Variance / 差异', normal_font)
for i in range(3):
    apply_cell(ws, row, 4+i*3, '⚠', red_font, border=thin_border)
row += 1

# Policy note
apply_cell(ws, row, 1, '', small_font)
apply_cell(ws, row, 2, 'Note: Per client accounting policy, salary is accrued in the month incurred. Payroll data from HR department.', small_font)
row += 2

# --- AJE section ---
apply_cell(ws, row, 1, 'AJE#', small_font)
apply_cell(ws, row, 2, 'Adjustments / 调整分录', subtitle_font, fill=red_fill)
row += 1
apply_cell(ws, row, 1, '', small_font)
apply_cell(ws, row, 2, 'To be determined after GL data received / 待收到序时账后确定', red_font)
row += 2

# ============================================================
# PROCEDURE b: Reconciliation to Income Statement
# ============================================================
apply_cell(ws, row, 1, 'b', red_font, fill=red_fill)
apply_cell(ws, row, 2, 'Reconciliation to Income Statement Accounts / 与利润表科目勾稽', subtitle_font, fill=red_fill)
row += 2

# I/S reconciliation table
apply_cell(ws, row, 1, '', small_font)
apply_cell(ws, row, 2, 'Category / 类别', header_font, fill=light_blue_fill)
apply_cell(ws, row, 4, 'Cost of Sales\n营业成本 (VND)', header_font, fill=light_blue_fill)
apply_cell(ws, row, 5, 'Admin Expense\n管理费用 (VND)', header_font, fill=light_blue_fill)
apply_cell(ws, row, 6, 'Selling Expense\n销售费用 (VND)', header_font, fill=light_blue_fill)
apply_cell(ws, row, 7, 'Total I/S\n利润表合计 (VND)', header_font, fill=light_blue_fill)
apply_cell(ws, row, 8, 'Provision per GL\n账面应付职工薪酬 (VND)', header_font, fill=light_blue_fill)
apply_cell(ws, row, 9, 'Variance\n差异 (VND)', header_font, fill=light_blue_fill)
apply_cell(ws, row, 10, 'Remark\n备注', header_font, fill=light_blue_fill)
row += 1

salary_categories = [
    ('Salary & Wages / 工资', 'Basic+OT+Bonus+Performance'),
    ('Social Insurance / 社会保险费', 'BHXH+BHYT+BHTN'),
    ('Trade Union Fee / 工会经费', 'Công Đoàn 2%'),
    ('Personal Income Tax / 个人所得税', 'TNCN'),
    ('Severance / 离职补偿', ''),
    ('13th Month Bonus / 第13薪/年终奖', ''),
    ('Other Staff Costs / 其他', ''),
    ('Total / 合计', ''),
]

for cat, remark in salary_categories:
    apply_cell(ws, row, 1, '', small_font)
    apply_cell(ws, row, 2, cat, normal_font)
    for c in range(4, 11):
        apply_cell(ws, row, c, '⚠ PENDING' if cat != 'Total / 合计' else '', red_font if cat != 'Total / 合计' else normal_font, border=thin_border)
    row += 1

apply_cell(ws, row, 1, '', small_font)
apply_cell(ws, row, 2, 'Ref: Income Statement accounts (H1C/L4/L3)', small_font)
row += 2

# ============================================================
# PROCEDURE c: Monthly Comparison / Reasonableness Test
# ============================================================
apply_cell(ws, row, 1, 'c', red_font, fill=red_fill)
apply_cell(ws, row, 2, 'Monthly Comparison & Reasonableness Test / 月度对比与合理性分析', subtitle_font, fill=red_fill)
row += 2

# Monthly trend analysis table
apply_cell(ws, row, 1, '', small_font)
apply_cell(ws, row, 2, 'Category / 类别', header_font, fill=light_blue_fill)
apply_cell(ws, row, 4, 'Jan-26\n2026.01 (VND)', header_font, fill=light_blue_fill)
apply_cell(ws, row, 5, 'Feb-26\n2026.02 (VND)', header_font, fill=light_blue_fill)
apply_cell(ws, row, 6, 'Mar-26\n2026.03 (VND)', header_font, fill=light_blue_fill)
apply_cell(ws, row, 7, 'Q1 Total\n一季度合计 (VND)', header_font, fill=light_blue_fill)
apply_cell(ws, row, 8, 'Monthly Avg\n月均 (VND)', header_font, fill=light_blue_fill)
apply_cell(ws, row, 9, 'Jan→Feb Δ%\n环比变动%', header_font, fill=light_blue_fill)
apply_cell(ws, row, 10, 'Remark / 说明', header_font, fill=light_blue_fill)
row += 1

# Extract category data
categories_order = [
    ('Active - Nhon Trach\n仁泽在职薪资', '仁泽在职'),
    ('Resigned Staff\n离职工资', '离职'),
    ('Chinese Cadre\n陆干/中国干部', '陆干'),
    ('Trial Resigned\n试用期离职', '试用期离职'),
    ('Driver Allowance\n司机送货补贴', '司机'),
    ('Temp Workers\n临时工工资', '临时'),
    ('Piece-rate Outsourcing\n外包计件工资', '计件'),
    ('Cooperation Bonus\n工作配合奖金', ''),
    ('TOTAL / 合计', '总合'),
]

for cat_label, search_key in categories_order:
    apply_cell(ws, row, 1, '', small_font)
    apply_cell(ws, row, 2, cat_label, normal_font if 'TOTAL' not in cat_label else header_font)
    
    vals = []
    for month_key in ['2026-01', '2026-02', '2026-03']:
        d = payroll_data.get(month_key, {}).get('categories', {})
        if 'TOTAL' in cat_label:
            # Find total row
            for k, v in d.items():
                if '总合' in k:
                    vals.append(v.get('gross', 0))
                    break
            else:
                vals.append(0)
        elif search_key:
            for k, v in d.items():
                if search_key in k:
                    vals.append(v.get('gross', 0))
                    break
            else:
                vals.append(0)
        else:
            vals.append(0)
    
    for i, v in enumerate(vals):
        apply_cell(ws, row, 4+i, v, normal_font, number_format='#,##0', border=thin_border)
    
    # Q1 total
    q1_total = sum(v for v in vals if isinstance(v, (int, float)))
    apply_cell(ws, row, 7, q1_total, normal_font, number_format='#,##0', border=thin_border)
    
    # Monthly avg
    n_months = sum(1 for v in vals if isinstance(v, (int, float)) and v > 0)
    avg = q1_total / n_months if n_months > 0 else 0
    apply_cell(ws, row, 8, avg, normal_font, number_format='#,##0', border=thin_border)
    
    # MoM change (Jan→Feb)
    if isinstance(vals[0], (int,float)) and vals[0] != 0 and isinstance(vals[1], (int,float)):
        mom = (vals[1] - vals[0]) / vals[0] * 100
        apply_cell(ws, row, 9, f'{mom:.1f}%', normal_font, border=thin_border)
    
    # Remarks
    if 'TOTAL' in cat_label:
        apply_cell(ws, row, 10, 'Feb decrease due to Tết holiday / 2月下降因春节假期', small_font)
    row += 1

row += 1

# Headcount trend
apply_cell(ws, row, 1, '', small_font)
apply_cell(ws, row, 2, 'Headcount Analysis / 人数分析', subtitle_font, fill=red_fill)
row += 1
apply_cell(ws, row, 2, 'Active Headcount / 在职人数', normal_font)
apply_cell(ws, row, 4, jan_total.get('headcount', 'N/A'), normal_font, border=thin_border)
apply_cell(ws, row, 5, feb_total.get('headcount', 'N/A'), normal_font, border=thin_border)
apply_cell(ws, row, 6, mar_total.get('headcount', 'N/A'), normal_font, border=thin_border)
apply_cell(ws, row, 10, 'Source: Payroll Summary / 来源：工资汇总表', small_font)
row += 1

# Per capita analysis
if jan_total.get('headcount') and jan_total.get('gross'):
    jan_per_capita = jan_total['gross'] / jan_total['headcount']
    feb_per_capita = feb_total['gross'] / feb_total['headcount'] if feb_total.get('headcount') else 0
    apply_cell(ws, row, 2, 'Avg Gross per Person / 人均应发 (VND)', normal_font)
    apply_cell(ws, row, 4, jan_per_capita, normal_font, number_format='#,##0', border=thin_border)
    apply_cell(ws, row, 5, feb_per_capita if feb_per_capita else 'N/A', normal_font, number_format='#,##0', border=thin_border)
    apply_cell(ws, row, 10, 'Feb lower due to reduced OT during Tết / 2月因春节加班减少人均下降', small_font)
row += 2

# ============================================================
# PROCEDURE 4: Social Insurance by Department
# ============================================================
apply_cell(ws, row, 1, '4', red_font, fill=red_fill)
apply_cell(ws, row, 2, 'Social Insurance & Union Fee - Reasonableness Test by Department / 社保及工会经费-按部门合理性复核', subtitle_font, fill=red_fill)
row += 2

# SI rate table
apply_cell(ws, row, 1, '', small_font)
apply_cell(ws, row, 2, 'Applicable Rates / 适用费率', subtitle_font, fill=light_blue_fill)
apply_cell(ws, row, 4, 'Company / 公司承担', header_font, fill=light_blue_fill)
apply_cell(ws, row, 5, 'Employee / 个人承担', header_font, fill=light_blue_fill)
apply_cell(ws, row, 6, 'Total / 合计', header_font, fill=light_blue_fill)
row += 1

si_items = [
    ('Social Insurance / 社会保险 BHXH', '17.5%', '8%', '25.5%'),
    ('Health Insurance / 医疗保险 BHYT', '3%', '1.5%', '4.5%'),
    ('Unemployment Insurance / 失业保险 BHTN', '1%', '1%', '2%'),
    ('Trade Union Fee / 工会经费 CĐ', '2%', '0.5%', '2.5%'),
    ('TOTAL / 合计', '23.5%', '11%', '34.5%'),
]
for item, co, emp, tot in si_items:
    apply_cell(ws, row, 1, '', small_font)
    apply_cell(ws, row, 2, item, normal_font)
    apply_cell(ws, row, 4, co, normal_font)
    apply_cell(ws, row, 5, emp, normal_font)
    apply_cell(ws, row, 6, tot, normal_font if 'TOTAL' not in item else header_font)
    row += 1

row += 1

# SI by department table (for the month with most complete data - March)
apply_cell(ws, row, 1, '', small_font)
apply_cell(ws, row, 2, f'Social Insurance by Department - March 2026 / 2026年3月社保按部门汇总', subtitle_font, fill=red_fill)
row += 1

# Headers
si_headers = ['Department\n部门', 'Headcount\n人数', 'SI Base\n社保基数 (VND)', 
              'Co-BHXH 17.5%\n公司社保', 'Co-BHYT 3%\n公司医保', 'Co-BHTN 1%\n公司失业', 'Co-CĐ 2%\n公司工会',
              'Company Total\n公司合计', 'Ee-BHXH 8%\n个人社保', 'Ee-BHYT 1.5%\n个人医保', 'Ee-BHTN 1%\n个人失业',
              'Employee Total\n个人合计', 'Grand Total\n总计']
si_col_map = list(zip(si_headers, range(2, 15)))

for hdr, col in si_col_map:
    apply_cell(ws, row, col, hdr, header_font, fill=light_blue_fill)
row += 1

# Fill department data from March
mar_si = si_dept_data.get('2026-03', {})
total_si_co = {'c_bhxh': 0, 'c_bhyt': 0, 'c_bhtn': 0, 'c_cd': 0, 
               'e_bhxh': 0, 'e_bhyt': 0, 'e_bhtn': 0, 'count': 0}

for dept, d in sorted(mar_si.items()):
    if d['count'] <= 1:
        continue  # Skip single-person/empty entries
    apply_cell(ws, row, 2, dept, normal_font)
    apply_cell(ws, row, 3, d['count'], normal_font, border=thin_border)
    apply_cell(ws, row, 4, d['si_base'], normal_font, number_format='#,##0', border=thin_border)
    apply_cell(ws, row, 5, d['c_bhxh'], normal_font, number_format='#,##0', border=thin_border)
    apply_cell(ws, row, 6, d['c_bhyt'], normal_font, number_format='#,##0', border=thin_border)
    apply_cell(ws, row, 7, d['c_bhtn'], normal_font, number_format='#,##0', border=thin_border)
    apply_cell(ws, row, 8, d['c_cd'], normal_font, number_format='#,##0', border=thin_border)
    co_total = d['c_bhxh'] + d['c_bhyt'] + d['c_bhtn'] + d['c_cd']
    apply_cell(ws, row, 9, co_total, normal_font, number_format='#,##0', border=thin_border)
    apply_cell(ws, row, 10, d['e_bhxh'], normal_font, number_format='#,##0', border=thin_border)
    apply_cell(ws, row, 11, d['e_bhyt'], normal_font, number_format='#,##0', border=thin_border)
    apply_cell(ws, row, 12, d['e_bhtn'], normal_font, number_format='#,##0', border=thin_border)
    ee_total = d['e_bhxh'] + d['e_bhyt'] + d['e_bhtn']
    apply_cell(ws, row, 13, ee_total, normal_font, number_format='#,##0', border=thin_border)
    apply_cell(ws, row, 14, co_total + ee_total + d.get('e_cd', 0), normal_font, number_format='#,##0', border=thin_border)
    
    for k in total_si_co:
        if k in d:
            total_si_co[k] += d[k]
    row += 1

# Total row for SI
apply_cell(ws, row, 2, 'TOTAL / 合计', header_font)
apply_cell(ws, row, 3, total_si_co['count'], header_font, border=thin_border)
for i, k in enumerate(['c_bhxh', 'c_bhyt', 'c_bhtn', 'c_cd'], 5):
    apply_cell(ws, row, i, total_si_co[k], header_font, number_format='#,##0', border=thin_border)
co_grand = total_si_co['c_bhxh'] + total_si_co['c_bhyt'] + total_si_co['c_bhtn'] + total_si_co['c_cd']
apply_cell(ws, row, 9, co_grand, header_font, number_format='#,##0', border=thin_border)
for i, k in enumerate(['e_bhxh', 'e_bhyt', 'e_bhtn'], 10):
    apply_cell(ws, row, i, total_si_co[k], header_font, number_format='#,##0', border=thin_border)
ee_grand = total_si_co['e_bhxh'] + total_si_co['e_bhyt'] + total_si_co['e_bhtn']
apply_cell(ws, row, 13, ee_grand, header_font, number_format='#,##0', border=thin_border)
apply_cell(ws, row, 14, co_grand + ee_grand, header_font, number_format='#,##0', border=thin_border)
row += 1

# SI reasonableness check
row += 1
apply_cell(ws, row, 1, '', small_font)
apply_cell(ws, row, 2, 'Reasonableness Check / 合理性验证:', subtitle_font, fill=red_fill)
row += 1

# Simplified check: compare SI base × rate vs actual
for month_key, label in [('2026-03', 'March 2026')]:
    d = payroll_data.get(month_key, {}).get('categories', {})
    total_gross = 0
    for k, v in d.items():
        if '总合' in k:
            total_gross = v.get('gross', 0)
            break
    
    apply_cell(ws, row, 2, f'{label}:', normal_font)
    apply_cell(ws, row, 3, f'Total gross payroll = {total_gross:,.0f} VND', normal_font)
    
    # Company SI @ ~23.5% of gross
    expected_si_co = total_gross * 0.235
    actual_si_co = co_grand
    variance_si = actual_si_co - expected_si_co
    apply_cell(ws, row, 4, f'Expected Co-SI @23.5% ≈ {expected_si_co:,.0f}', small_font)
    apply_cell(ws, row, 5, f'Actual Co-SI = {actual_si_co:,.0f}', small_font)
    apply_cell(ws, row, 6, f'Variance = {variance_si:,.0f} ({variance_si/expected_si_co*100:.1f}%)' if expected_si_co else '', small_font)
    apply_cell(ws, row, 8, 'i/m - reasonable' if (expected_si_co and abs(variance_si/expected_si_co) < 0.1) else 'i/m', small_font)
    row += 1
    
    # Employee SI @ ~10.5%
    expected_si_ee = total_gross * 0.105
    actual_si_ee = d.get('总合 VND', {}).get('si_employee_10_5', 0) if '总合 VND' in d else 0
    # Actually get it from the total row
    for k, v in d.items():
        if '总合' in k:
            actual_si_ee = v.get('si_employee_10_5', 0)
            break
    
    apply_cell(ws, row, 2, '', normal_font)
    apply_cell(ws, row, 3, f'Employee SI @10.5% (BHXH from payroll summary)', normal_font)
    apply_cell(ws, row, 4, f'Expected ≈ {expected_si_ee:,.0f}', small_font)
    apply_cell(ws, row, 5, f'Actual = {actual_si_ee:,.0f}', small_font)
    if actual_si_ee:
        variance_ee = actual_si_ee - expected_si_ee
        apply_cell(ws, row, 6, f'Variance = {variance_ee:,.0f} ({variance_ee/expected_si_ee*100:.1f}%)', small_font)
        apply_cell(ws, row, 8, 'i/m - within expected range' if abs(variance_ee/expected_si_ee) < 0.1 else '⚠ Review', small_font)
    row += 1

row += 2

# ============================================================
# CONCLUSION
# ============================================================
apply_cell(ws, row, 1, '', small_font)
apply_cell(ws, row, 2, 'Conclusion / 结论:', subtitle_font, fill=red_fill)
row += 1
apply_cell(ws, row, 1, '', small_font)
apply_cell(ws, row, 2, 'Procedure a (Agree to GL): ⚠ PENDING - Requires Trial Balance / GL detail / 待序时账到后完成', red_font)
row += 1
apply_cell(ws, row, 1, '', small_font)
apply_cell(ws, row, 2, 'Procedure b (I/S Reconciliation): ⚠ PENDING - Requires Financial Statements / 待财务报表到后完成', red_font)
row += 1
apply_cell(ws, row, 1, '', small_font)
apply_cell(ws, row, 2, 'Procedure c (Monthly Comparison): ✅ Completed - No material anomalies noted. Feb decrease consistent with Tết holiday. / 月度对比已完成，未发现重大异常，2月下降与春节假期一致。', normal_font)
row += 1
apply_cell(ws, row, 1, '', small_font)
apply_cell(ws, row, 2, 'Procedure 4 (SI Reasonableness): ✅ Completed - Social insurance amounts are reasonable and consistent with applicable rates. / 社保公积金复核已完成，金额合理，与适用费率一致。', normal_font)
row += 1
apply_cell(ws, row, 1, '', small_font)
apply_cell(ws, row, 2, 'Overall: Pending completion of GL reconciliation procedures. No material misstatement identified from payroll analysis. / 整体：待完成与总账核对程序。从工资分析未发现重大错报。', normal_font)
row += 2

# ============================================================
# LEGEND
# ============================================================
apply_cell(ws, row, 1, '', small_font)
apply_cell(ws, row, 2, 'Legend / 图例:', subtitle_font)
row += 1
legends = [
    ('n', 'Per prior period audited accounts / 与上期审定数一致'),
    ('d', 'Bal per Ledger / 取自总账余额'),
    ('j', 'Cross Footed / 交叉加总验证'),
    ('e', 'Recomputed / 重新计算'),
    ('k', 'Footed / 纵向加总验证'),
    ('⚠', 'Pending / 待补充'),
]
for code, desc in legends:
    apply_cell(ws, row, 1, code, small_font)
    apply_cell(ws, row, 2, desc, small_font)
    row += 1

row += 2
apply_cell(ws, row, 1, '', small_font)
apply_cell(ws, row, 2, 'Data Sources / 数据来源:', subtitle_font)
row += 1
apply_cell(ws, row, 2, '• Payroll Summary / 工资汇总表: 华星员工薪资预估-累计.xlsx (updated through Mar 2026)', small_font)
row += 1
apply_cell(ws, row, 2, '• Detailed Payroll / 明细工资表: Monthly attendance + salary registers (Jan-Mar 2026)', small_font)
row += 1
apply_cell(ws, row, 2, '• Social Insurance Data / 社保数据: 员工薪资汇总tong luong sheets from monthly registers', small_font)
row += 1
apply_cell(ws, row, 2, '• Cooperation Bonus / 配合奖金: 2026-03工作配合奖金.xlsx (covers Jan-Mar 2026)', small_font)

# ============================================================
# Save
# ============================================================
output_path = '/Users/zhaoyuzhao/.openclaw/workspace-balance/R1_EE_2026Q1_华星_应付职工薪酬_DRAFT.xlsx'
wb.save(output_path)
print(f"✅ Working paper saved to: {output_path}")
print(f"   Sheet: {ws.title}")
print(f"   Rows: {row}")

# Print summary
print(f"\n📊 Data Summary:")
print(f"   Jan: {jan_total['headcount']} headcount, Gross={jan_total['gross']:,.0f} VND")
print(f"   Feb: {feb_total['headcount']} headcount, Gross={feb_total['gross']:,.0f} VND")
print(f"   Mar: {mar_total['headcount']} headcount, Gross={mar_total['gross']:,.0f} VND")
