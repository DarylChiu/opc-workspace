#!/usr/bin/env python3
"""
tokhai_extract.py — v2.6 ToKhai 内容级抽取模块
职责：从单个 ToKhai .xls/.xlsx 文件读内容，判定是否「已通关」，
      抽取 关单号 / 报关日期 / 发票号 / 发票金额（供方案B金额勾稽用）。
判定完全基于文件内容，不依赖文件名（文件名带不带QDTQ只作辅助线索）。
"""
import re
import os
import unicodedata

try:
    import xlrd  # .xls
except ImportError:
    xlrd = None
try:
    import openpyxl  # .xlsx
except ImportError:
    openpyxl = None


def _norm(s):
    """去越南语声调 + 小写，用于关键词匹配
    注：越南语 đ/Đ 是独立字母，NFD 不会分解，需手动映射为 d"""
    s = unicodedata.normalize('NFD', str(s))
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn').lower()
    return s.replace('đ', 'd').replace('Đ', 'd')


def _parse_amount(token):
    """越南数字格式 '9.076,22' / '35.153,58' → float"""
    if token is None:
        return None
    num = str(token).strip()
    m = re.search(r'[\d.,]+', num)
    if not m:
        return None
    num = m.group(0)
    if '.' in num and ',' in num:          # 9.076,22
        num = num.replace('.', '').replace(',', '.')
    elif ',' in num:                        # 9076,22
        num = num.replace(',', '.')
    else:
        # 只有点：可能是千分位 9.076 或小数 9.07
        parts = num.split('.')
        if len(parts) > 1 and all(len(x) == 3 for x in parts[1:]):
            num = num.replace('.', '')      # 千分位
    try:
        return float(num)
    except ValueError:
        return None


def _load_rows(filepath):
    """读第一个 sheet(TKN) 的所有行，返回 list[list[str]]"""
    ext = os.path.splitext(filepath)[1].lower()
    rows = []
    if ext == '.xls':
        if xlrd is None:
            raise RuntimeError('xlrd 未安装，无法读 .xls')
        wb = xlrd.open_workbook(filepath)
        sh = wb.sheet_by_index(0)
        for r in range(sh.nrows):
            rows.append([str(sh.cell_value(r, c)).strip() for c in range(sh.ncols)])
    elif ext in ('.xlsx', '.xlsm'):
        if openpyxl is None:
            raise RuntimeError('openpyxl 未安装，无法读 .xlsx')
        wb = openpyxl.load_workbook(filepath, data_only=True)
        sh = wb[wb.sheetnames[0]]
        for row in sh.iter_rows(values_only=True):
            rows.append([str(c).strip() if c is not None else '' for c in row])
        wb.close()
    else:
        raise RuntimeError(f'不支持的格式: {ext}')
    return rows


# 状态标题判定关键词（内容级）
STATUS_THONGQUAN = 'thong quan'                 # (thông quan) = 已通关
STATUS_PHANLUONG = 'thong bao ket qua phan luong'  # 分流通知 = 中间态
STATUS_DRAFT = ['ban xac nhan noi dung', 'in thu']  # 内容确认书/试打 = 草稿


def extract_tokhai(filepath):
    """
    返回 dict:
      is_customs_doc : 是否是报关单文件（TKN sheet 结构）
      status_text    : 状态标题原文
      status         : 'THONG_QUAN' | 'PHAN_LUONG' | 'DRAFT' | 'OTHER'
      is_cleared     : 是否已通关 (status == THONG_QUAN)
      so_to_khai     : 关单号
      ngay_dang_ky   : 报关日期(原始 DD/MM/YYYY)
      ngay_ymd       : 报关日期(YYYY/MM/DD)
      so_hoa_don     : 发票号(报关单内记载)
      amount         : 发票金额(Tổng trị giá hóa đơn)
      error          : 读取错误信息
    """
    res = {
        'file': os.path.basename(filepath),
        'is_customs_doc': False, 'status_text': '', 'status': 'OTHER',
        'is_cleared': False, 'so_to_khai': '', 'ngay_dang_ky': '',
        'ngay_ymd': '', 'so_hoa_don': '', 'amount': None, 'hs_code': '', 'error': '',
    }
    try:
        rows = _load_rows(filepath)
    except Exception as e:
        res['error'] = str(e)
        return res

    for idx, cells in enumerate(rows):
        nz = [v for v in cells if v]
        if not nz:
            continue
        line = ' | '.join(nz)
        n = _norm(line)

        # 状态标题：通常在前 3 行含 "to khai hang hoa nhap khau"
        if 'to khai hang hoa nhap khau' in n or 'ban xac nhan noi dung' in n:
            if not res['status_text']:
                res['status_text'] = line
                res['is_customs_doc'] = True
                if STATUS_THONGQUAN in n:
                    res['status'] = 'THONG_QUAN'; res['is_cleared'] = True
                elif STATUS_PHANLUONG in n:
                    res['status'] = 'PHAN_LUONG'
                elif any(k in n for k in STATUS_DRAFT):
                    res['status'] = 'DRAFT'

        # 关单号 Số tờ khai
        if n.startswith('so to khai') and not res['so_to_khai']:
            m = re.search(r'(\d{10,13})', line)
            if m:
                res['so_to_khai'] = m.group(1)

        # 报关日期 Ngày đăng ký
        if 'ngay dang ky' in n and not res['ngay_dang_ky']:
            m = re.search(r'(\d{2}/\d{2}/\d{4})', line)
            if m:
                res['ngay_dang_ky'] = m.group(1)
                d, mo, y = m.group(1).split('/')
                res['ngay_ymd'] = f'{y}/{mo}/{d}'

        # 发票号 Số hóa đơn
        if n.startswith('so hoa don') and not res['so_hoa_don']:
            # 形如 "Số hóa đơn | A - HMRFZ220134"
            parts = [p.strip() for p in line.split('|')]
            if len(parts) > 1:
                val = parts[-1]
                val = re.sub(r'^[A-Z]\s*-\s*', '', val).strip()
                res['so_hoa_don'] = val

        # 金额 Tổng trị giá hóa đơn（排除 tính thuế / phân bổ 行）
        if 'tong tri gia hoa don' in n and res['amount'] is None:
            res['amount'] = _parse_amount(nz[-1])

        # 代表性HS编码 Mã số hàng hóa đại diện của tờ khai
        if 'ma so hang hoa dai dien' in n and not res['hs_code']:
            m = re.search(r'ma so hang hoa dai dien[^\d]*(\d{4,8})', n)
            if m:
                res['hs_code'] = m.group(1)

    return res


def is_tokhai_filename(fname):
    """辅助线索：文件名疑似 ToKhai（用于初筛候选，非判据）"""
    u = fname.upper()
    return ('TOKHAI' in u or 'QDTQ' in u) and fname.lower().endswith(('.xls', '.xlsx', '.xlsm'))


if __name__ == '__main__':
    import sys
    r = extract_tokhai(sys.argv[1])
    for k, v in r.items():
        print(f'{k:15}: {v}')
