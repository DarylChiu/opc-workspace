#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""误餐补贴明细表 v0.5.3：按征集意见更新 标准与参数 + 填写说明 + 保护参数修复"""
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
import shutil

SRC = 'reports/附件-误餐补贴申请明细表-v0.5.2-20260805.xlsx'
DST = 'reports/附件-误餐补贴申请明细表-v0.5.3-20260805.xlsx'
shutil.copy(SRC, DST)
PWD = "13226469877"

wb = openpyxl.load_workbook(DST)

# ═══ 1. 标准与参数：城市等级对照 K/L 列 新增二类城市 ═══
ws = wb['标准与参数']
# 现有 K2:L29 (K29=其他三类城市)。新增行 K30:L36
new_cities = [
    ('Thủ Đức（守德郡）', '二类城市'),
    ('An Giang（安江）', '二类城市'),
    ('Vĩnh Long（永隆）', '二类城市'),
    ('Bến Tre（槟知）', '二类城市'),
    ('Tiền Giang（前江）', '二类城市'),
    ('Nam Định（南定）', '二类城市'),
    ('Ninh Bình（宁平）', '二类城市'),
]
r = 30
for city, grade in new_cities:
    ws.cell(row=r, column=11, value=city)   # K
    ws.cell(row=r, column=12, value=grade)  # L
    ws.cell(row=r, column=11).alignment = Alignment(vertical='center')
    ws.cell(row=r, column=12).alignment = Alignment(vertical='center')
    r += 1
# 样式与现有表一致（简单边框字体）
thin = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'))
for rr in range(30, 37):
    for cc in (11, 12):
        ws.cell(row=rr, column=cc).border = thin
        ws.cell(row=rr, column=cc).font = Font(size=10)
# K1 标题更新说明
ws.cell(row=1, column=11, value='城市等级对照（附件2·Nghị quyết 111/2025）※2026-08-05征集意见新增二类城市：守德/安江/永隆/槟知/前江/南定/宁平')

# ═══ 2. 申请明细：A2 填写说明 补充误餐场景规则 ═══
ws2 = wb['申请明细']
old_a2 = ws2['A2'].value
add_note = (
    '｜⚠️ 误餐规则（2026-08-05征集意见定稿）：'
    '①当日已报销"业务招待费-餐费"的餐次，不再给予对应餐次误餐补贴（"是否招待"列须如实勾选）；'
    '②误餐时间须按实际门禁进出记录填写，严禁虚报；财务部审核将抽检门禁记录，虚报按管理办法第6章处理；'
    '③当日出差仅报销误餐补贴的，报销时仅需出差审批记录，不需出入打卡证明；'
    '④返程过晚直接回家未回公司签退的，以部门总监签字确认为准（本表签字栏）。'
)
ws2['A2'] = str(old_a2) + add_note

# ═══ 3. 合规声明 A22 补充门禁真实性 ═══
old_a22 = ws2['A22'].value
if old_a22:
    ws2['A22'] = str(old_a22) + ' ②误餐时间均依据实际门禁进出记录填写，如有虚报愿接受公司处理。'

# ═══ 4. 保护参数修复：允许选择单元格（只锁编辑）═══
for name in wb.sheetnames:
    w = wb[name]
    w.protection.sheet = True
    w.protection.password = PWD
    w.protection.selectLockedCells = True
    w.protection.selectUnlockedCells = True
    w.protection.objects = False
    w.protection.scenarios = False

wb.save(DST)
print('✅ v0.5.3 saved:', DST)
