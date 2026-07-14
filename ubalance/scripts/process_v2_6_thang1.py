#!/usr/bin/env python3
"""
process_v2_6_thang1.py — v2.6 THÁNG 1-2023 完整样张
13列 = v2.5原10列 + 新增(关单号/报关日期/ToKhai勾稽)
列序: 序号|发票号|关单号|报关日期|金额(USD)|日期|供应商|贷款用途分类|分类依据|置信度|三件套完整性|ToKhai勾稽|备注
按 STT 升序排序。内容级判定，直接扫 RM-Database/2023。
"""
import os, sys, re
sys.path.insert(0, os.path.dirname(__file__))
from tokhai_extract import extract_tokhai, is_tokhai_filename, _norm
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

try:
    import pdfplumber
    PDF_READY = True
except ImportError:
    PDF_READY = False

DISK = '/Volumes/TOSHIBA EXT/HUATEX贷款材料'
L1 = os.path.join(DISK, 'L1-Raw Materials-Year2023.xlsx')
MONTH_DIR = os.path.join(DISK, 'RM-Database/2023', 'THÁNG 1-2023')
WS_COPY = '/Users/zhaoyuzhao/.openclaw/workspace-balance'
TS = datetime.now().strftime('%Y%m%d_%H%M')
OUT = os.path.join(DISK, f'Outcome1-v2.6-THANG1-{TS}.xlsx')
OUT_WS = os.path.join(WS_COPY, f'Outcome1-v2.6-THANG1-{TS}.xlsx')
TOL = 0.001

# ---------- 内容分类器（精简自 v2.5，聚焦三件套判定）----------
def pdf_text(fp, pages=2):
    t = ''
    if not PDF_READY:
        return t
    try:
        with pdfplumber.open(fp) as pdf:
            for p in pdf.pages[:pages]:
                x = p.extract_text()
                if x: t += x + '\n'
    except Exception:
        pass
    return t.strip()

def xls_text(fp):
    t = ''
    try:
        if fp.lower().endswith('.xls'):
            import xlrd
            wb = xlrd.open_workbook(fp)
            for sh in wb.sheets()[:3]:
                for r in range(min(sh.nrows, 60)):
                    t += ' '.join(str(sh.cell_value(r,c)) for c in range(sh.ncols)) + '\n'
        else:
            wb = openpyxl.load_workbook(fp, data_only=True)
            for sn in wb.sheetnames[:3]:
                for i, row in enumerate(wb[sn].iter_rows(values_only=True)):
                    if i > 60: break
                    t += ' '.join(str(c) for c in row if c is not None) + '\n'
            wb.close()
    except Exception:
        pass
    return t.strip()

def classify(fp):
    """返回 (category, evidence). 保留类: INVOICE_PDF/TOKHAI_QDTQ/BL/SWB/AWB"""
    fname = os.path.basename(fp)
    ext = os.path.splitext(fp)[1].lower()
    if ext in ('.xls', '.xlsx', '.xlsm'):
        text = xls_text(fp)
        tu = text.upper()
        n = _norm(text)
        # ToKhai：内容级判定通关
        if 'to khai hang hoa nhap khau' in n or is_tokhai_filename(fname):
            r = extract_tokhai(fp)
            if r['is_cleared']:
                return 'TOKHAI_QDTQ', '已通关报关单(thông quan)'
            return 'TOKHAI_OTHER', f'报关单未通关({r["status"]})'
        if any(k in tu for k in ['SALES CONFIRMATION','COMMERCIAL INVOICE','PACKING LIST','PROFORMA INVOICE']):
            return 'INVOICE_EXCEL', '清关资料Excel版'
        return 'UNKNOWN', 'Excel内容不明'
    if ext == '.pdf':
        text = pdf_text(fp); tu = text.upper()
        if 'BILL OF LADING' in tu:
            return 'BL', '正本提单'
        if 'SEA WAYBILL' in tu or 'NON-NEGOTIABLE' in tu:
            return 'SWB', '海运单'
        if any(k in tu for k in ['AIR WAYBILL','HAWB','MAWB','AWB NO']):
            return 'AWB', '空运提单'
        if 'TELEX RELEASE' in tu or 'SURRENDER' in tu:
            return 'BL', '电放提单'
        if any(k in tu for k in ['SALES CONFIRMATION','COMMERCIAL INVOICE','PACKING LIST','PROFORMA INVOICE']):
            return 'INVOICE_PDF', '销售合同/商业发票/装箱单'
        if any(k in tu for k in ['CERTIFICATE OF ORIGIN','FORM E']):
            return 'CO_CERT', '产地证'
        has_ship = any(k in tu for k in ['SHIPPER','CONSIGNEE'])
        has_vessel = any(k in tu for k in ['VESSEL','VOYAGE','CONTAINER','PORT OF LOADING'])
        if has_ship and has_vessel:
            return 'BL', '海运文件(特征匹配)'
        # 文件名兜底
        u = fname.upper()
        if 'BL' in u or 'SWB' in u: return 'BL', '文件名判定:提单'
        if 'AWB' in u or 'HAWB' in u: return 'AWB', '文件名判定:空运单'
        if '清关资料' in fname or '中英文明细' in fname: return 'INVOICE_PDF', '文件名判定:清关资料PDF'
        return 'UNKNOWN', f'无法判定({len(text)}字符)'
    return 'UNKNOWN', f'不支持格式{ext}'

EQUIP_KW = ['设备','机器','验布机','水洗机','染色机','经编机','织机','一体机','配件','钢筘',
            '盘头','料框','储油槽','锅炉','立库','A字架','FR柜','人脸识别','喷淋塔','试验机']
RAW_KW = ['胚布','成品','纱线','助剂','染料','毛坯','布']

def classify_equipment(text):
    text = text.replace('\xa0',' ')
    e = [k for k in EQUIP_KW if re.search(k, text, re.I)]
    r = [k for k in RAW_KW if re.search(k, text, re.I)]
    if e and r: return '混合(待拆分)', f'设备:{",".join(e)}+原料:{",".join(r)}', '中'
    if e: return '设备', ",".join(e), '高'
    if r: return '原材料', ",".join(r), '高'
    return '⚠️待确认', '无关键词命中', '低'

# ---------- 读 L1 (col1=STT,2=供应商,3=发票号,4=关单号,5=日期,6=金额) ----------
wb = openpyxl.load_workbook(L1, data_only=True)
ws = wb.active
l1_rows, l1_map = [], {}
for row in ws.iter_rows(min_row=4, values_only=True):
    if not row or len(row) < 6 or not row[2]: continue
    inv = str(row[2]).strip()
    if inv == 'None': continue
    try: amt = float(row[5]) if row[5] is not None else None
    except (ValueError, TypeError): amt = None
    d = row[4]
    if isinstance(d, datetime): dstr = d.strftime('%Y/%m/%d')
    else: dstr = str(d).replace('-','/').split(' ')[0] if d else ''
    rec = {'stt': row[0], 'supplier': str(row[1]).strip() if row[1] else '',
           'inv': inv, 'l1_date': dstr, 'amount': amt}
    l1_rows.append(rec); l1_map[inv] = rec
wb.close()
l1_invoices = set(l1_map.keys())

def resolve_invoice(dirname):
    core = re.sub(r'^\s*\d+\.\s*', '', dirname).strip()
    tokens = re.split(r'\s+', core)
    cands = [t for t in tokens if t in l1_invoices]
    if cands: return sorted(cands, key=len, reverse=True)[0]
    for t in tokens:
        if re.match(r'^[A-Z]{2,}.*\d', t.upper()): return t
    return tokens[0] if tokens else core

# ---------- 扫描 THÁNG 1 各目录 ----------
rows_out = {}   # inv -> full row dict
print(f'扫描 THÁNG 1 目录...')
for d in sorted(os.listdir(MONTH_DIR)):
    dp = os.path.join(MONTH_DIR, d)
    if not os.path.isdir(dp): continue
    primary_inv = resolve_invoice(d)
    files = [f for f in os.listdir(dp) if os.path.isfile(os.path.join(dp, f))]
    # 分类每个文件
    cats = []
    cleared_tokhai = []
    for f in files:
        fp = os.path.join(dp, f)
        cat, ev = classify(fp)
        cats.append(cat)
        if cat == 'TOKHAI_QDTQ':
            r = extract_tokhai(fp)
            if r['is_cleared']: cleared_tokhai.append(r)
    # 三件套完整性
    has_inv = any(c in ('INVOICE_PDF','TOKHAI_QDTQ') for c in cats)
    has_bl = any(c in ('BL','SWB','AWB') for c in cats)
    has_tokhai = 'TOKHAI_QDTQ' in cats
    missing = []
    if not has_inv: missing.append('Invoice')
    if not has_bl: missing.append('B/L')
    if not has_tokhai: missing.append('ToKhai')
    complete = '✅完整' if (has_inv and has_bl) else '⚠️缺'+','.join(missing)
    # 设备分类
    cls, src, cf = classify_equipment(d + ' ' + ' '.join(files))
    # ToKhai 勾稽（方案B：按金额归到 L1 发票，处理兄弟发票）
    tokhai_by_inv = {}
    for c in cleared_tokhai:
        for rec in l1_rows:
            if rec['amount'] is not None and c['amount'] is not None and abs(rec['amount']-c['amount'])<TOL:
                k = rec['inv']
                tokhai_by_inv.setdefault(k, []).append(c)
                break
    # 为该目录相关的每个发票建行
    inv_keys = set([primary_inv]) | set(tokhai_by_inv.keys())
    for inv in inv_keys:
        rec = l1_map.get(inv)
        tks = tokhai_by_inv.get(inv, [])
        if tks:
            so = ';'.join(t['so_to_khai'] for t in tks)
            date = ';'.join(t['ngay_ymd'] for t in tks)
            grade = '✅金额一致' if len(tks)==1 else f'⚠️分批报关{len(tks)}单'
            note = '' if len(tks)==1 else '同发票多张通关单,金额均符,已全列'
        else:
            so, date, grade = '', '', '❌未匹配'
            note = '⚠️无通关报关单-待人工复核' if inv in l1_invoices else '目录发票不在L1放款清单'
        rows_out[inv] = {
            'stt': rec['stt'] if rec else '', 'inv': inv, 'so': so, 'date': date,
            'amount': rec['amount'] if rec else None, 'l1_date': rec['l1_date'] if rec else '',
            'supplier': rec['supplier'] if rec else '',
            'cls': cls, 'src': src, 'cf': cf, 'complete': complete,
            'grade': grade, 'note': note,
        }

# ---------- 排序：STT 升序 ----------
def stt_key(r):
    try: return (0, int(r['stt']))
    except (ValueError, TypeError): return (1, 0)
ordered = sorted(rows_out.values(), key=stt_key)

# ---------- 写 Outcome1 (13列) ----------
wbo = openpyxl.Workbook(); sh = wbo.active
sh.title = '银行放款明细表'
headers = ['序号','发票号','关单号','报关日期','金额(USD)','日期','供应商',
           '贷款用途分类','分类依据','置信度','三件套完整性','ToKhai勾稽','备注']
hfill = PatternFill('solid', fgColor='4472C4')
hfont = Font(color='FFFFFF', bold=True, size=10)
thin = Side(style='thin', color='BFBFBF'); border = Border(thin,thin,thin,thin)
for c,h in enumerate(headers,1):
    cell=sh.cell(1,c,h); cell.fill=hfill; cell.font=hfont
    cell.alignment=Alignment(horizontal='center',vertical='center'); cell.border=border
r=2
for row in ordered:
    vals=[row['stt'],row['inv'],row['so'],row['date'],row['amount'],row['l1_date'],
          row['supplier'],row['cls'],row['src'],row['cf'],row['complete'],row['grade'],row['note']]
    for c,v in enumerate(vals,1):
        cell=sh.cell(r,c,v); cell.border=border
        cell.alignment=Alignment(horizontal='center' if c in(1,3,4,10) else 'left',vertical='center')
        if c==5 and isinstance(v,(int,float)): cell.number_format='#,##0.00'
    r+=1
widths=[7,17,15,13,13,12,24,14,20,8,18,14,26]
for c,w in enumerate(widths,1):
    sh.column_dimensions[openpyxl.utils.get_column_letter(c)].width=w
sh.freeze_panes='A2'
wbo.save(OUT)
import shutil
shutil.copy2(OUT, OUT_WS)
print(f'✅ 已生成: {OUT}')
print(f'✅ workspace副本: {OUT_WS}')
print(f'行数: {len(ordered)} | 列数: 13')
print(f'STT排序: {[r["stt"] for r in ordered]}')
