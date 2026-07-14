#!/usr/bin/env python3
"""
gen_outcome1_thang1.py — v2.6 THÁNG 1-2023 样张 Outcome1 生成
新增字段：发票号后加「关单号」+「报关日期」(方案B，从ToKhai内容独立抓取)
判定：内容级通关状态 + 金额完全相等勾稽 + 多通关分批备注 + 无通关单待复核
"""
import os, sys, re
sys.path.insert(0, os.path.dirname(__file__))
from tokhai_extract import extract_tokhai, is_tokhai_filename
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

DISK = '/Volumes/TOSHIBA EXT/HUATEX贷款材料'
L1 = os.path.join(DISK, 'L1-Raw Materials-Year2023.xlsx')
MONTH_DIR = os.path.join(DISK, 'RM-Database/2023', 'THÁNG 1-2023')
TS = datetime.now().strftime('%Y%m%d_%H%M')
OUT = os.path.join(DISK, f'Outcome1-v2.6-THANG1-样张-{TS}.xlsx')
TOL = 0.001

# ---- 读 L1（本月发票清单：发票号→金额/供应商）----
wb = openpyxl.load_workbook(L1, data_only=True)
ws = wb.active
l1_rows = []       # 保持 L1 顺序
l1_map = {}
for row in ws.iter_rows(min_row=4, values_only=True):
    if not row or len(row) < 7 or not row[2]:
        continue
    inv = str(row[2]).strip()
    if inv == 'None':
        continue
    try:
        amt = float(row[5]) if row[5] is not None else None
    except (ValueError, TypeError):
        amt = None
    rec = {'stt': row[0], 'supplier': str(row[1]).strip() if row[1] else '',
           'inv': inv, 'l1_so': str(row[3]).strip() if row[3] else '',
           'l1_date': row[4], 'amount': amt}
    l1_rows.append(rec)
    l1_map[inv] = rec
wb.close()
l1_invoices = set(l1_map.keys())

# ---- 目录名 → 发票号（扫描所有token，取能匹配L1的；跳过NIKE等噪音）----
def resolve_invoice(dirname):
    core = re.sub(r'^\s*\d+\.\s*', '', dirname).strip()
    tokens = re.split(r'[\s]+', core)
    # 先找精确匹配 L1 的 token（优先含后缀的长匹配）
    cands = [t for t in tokens if t in l1_invoices]
    if cands:
        return sorted(cands, key=len, reverse=True)[0]
    # 无精确匹配：返回第一个像发票号的 token（字母+数字）
    for t in tokens:
        if re.match(r'^[A-Z]{2,}.*\d', t.upper()):
            return t
    return tokens[0] if tokens else core

# ---- 遍历目录，抓 ToKhai ----
dir_index = {}   # invoice_no -> {so, date, note, amount}
for d in sorted(os.listdir(MONTH_DIR)):
    dp = os.path.join(MONTH_DIR, d)
    if not os.path.isdir(dp):
        continue
    inv = resolve_invoice(d)
    l1e = l1_map.get(inv)
    l1amt = l1e['amount'] if l1e else None
    cleared = []
    for f in os.listdir(dp):
        if is_tokhai_filename(f):
            r = extract_tokhai(os.path.join(dp, f))
            if r['is_cleared']:
                cleared.append(r)
    # 同目录可能含兄弟发票的多张通关单：按金额勾稽归到各自发票
    for c in cleared:
        # 找金额完全相等的 L1 发票
        for rec in l1_rows:
            if rec['amount'] is not None and c['amount'] is not None \
               and abs(rec['amount'] - c['amount']) < TOL:
                key = rec['inv']
                if key not in dir_index:
                    dir_index[key] = {'so': c['so_to_khai'], 'date': c['ngay_ymd'],
                                      'amount': c['amount'], 'note': '', 'count': 1,
                                      'src_dir': d}
                else:
                    # 已有：真·多通关同额 → 分批报关
                    dir_index[key]['so'] += f";{c['so_to_khai']}"
                    dir_index[key]['date'] += f";{c['ngay_ymd']}"
                    dir_index[key]['count'] += 1
                    dir_index[key]['note'] = f"⚠️分批报关{dir_index[key]['count']}单"
                break

# ---- 生成 Outcome1（列：序号|发票号|关单号|报关日期|金额|供应商|通关状态|备注）----
wbo = openpyxl.Workbook()
sh = wbo.active
sh.title = '银行放款明细表-THANG1样张'
headers = ['序号', '发票号', '关单号', '报关日期', '金额(USD)', '供应商', 'ToKhai勾稽', '备注']
hfill = PatternFill('solid', fgColor='1F4E78')
hfont = Font(color='FFFFFF', bold=True, size=11)
thin = Side(style='thin', color='BFBFBF')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
for c, h in enumerate(headers, 1):
    cell = sh.cell(1, c, h)
    cell.fill = hfill; cell.font = hfont
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

# 只输出 THÁNG 1 目录涉及的发票（样张范围）
thang1_invs_ordered = []
_seen = set()
for d in sorted(os.listdir(MONTH_DIR)):
    if not os.path.isdir(os.path.join(MONTH_DIR, d)):
        continue
    iv = resolve_invoice(d)
    if iv not in _seen:
        _seen.add(iv)
        thang1_invs_ordered.append(iv)
    # 兄弟发票：同目录金额勾稽出的其他发票也纳入
    for k in dir_index:
        if k not in _seen and dir_index[k].get('src_dir') == d:
            _seen.add(k)
            thang1_invs_ordered.append(k)

r = 2
matched = 0
for inv in thang1_invs_ordered:
    rec = l1_map.get(inv) or {'stt': '', 'supplier': '', 'inv': inv, 'amount': None}
    idx = dir_index.get(inv)
    if idx:
        so, date, note = idx['so'], idx['date'], idx['note']
        status = '✅金额一致'
        matched += 1
    else:
        so, date, note = '', '', '⚠️无通关报关单-待人工复核'
        status = '❌未匹配'
    vals = [rec['stt'], inv, so, date, rec['amount'], rec['supplier'], status, note]
    for c, v in enumerate(vals, 1):
        cell = sh.cell(r, c, v)
        cell.border = border
        cell.alignment = Alignment(horizontal='center' if c in (1,3,4,7) else 'left',
                                   vertical='center')
        if c == 5 and isinstance(v, (int, float)):
            cell.number_format = '#,##0.00'
    r += 1

# 列宽
widths = [8, 18, 15, 13, 14, 26, 12, 26]
for c, w in enumerate(widths, 1):
    sh.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w
sh.freeze_panes = 'A2'
wbo.save(OUT)

# 仅统计本月（THÁNG 1 目录里出现的）
thang1_invs = set()
for d in os.listdir(MONTH_DIR):
    if os.path.isdir(os.path.join(MONTH_DIR, d)):
        thang1_invs.add(resolve_invoice(d))
t1_matched = sum(1 for i in thang1_invs if i in dir_index)

print(f'✅ 样张已生成: {OUT}')
print(f'L1 发票总数(全2023): {len(l1_rows)}')
print(f'THÁNG 1 目录发票数: {len(thang1_invs)}')
print(f'THÁNG 1 抓到通关单: {t1_matched}')
print(f'全表已抓关单号的发票行: {matched}')
