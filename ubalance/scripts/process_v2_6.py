#!/usr/bin/env python3
"""
process_v2_6.py — v2.6 贷款材料整理(2023 HUATEX批次)
铁律1: 严格从 L1 出发 —— 只遍历 L1 发票行，绝不新增 L1 以外的行
铁律2: 内容级判定 —— ToKhai读通关状态/金额/HS码，Invoice读品名，不靠文件名
13列: 序号|发票号|关单号|报关日期|金额(USD)|日期|供应商|贷款用途分类|分类依据|置信度|三件套完整性|ToKhai勾稽|备注
用法: python3 process_v2_6.py [月份目录名，默认全部]  例: "THÁNG 1-2023"
"""
import os, sys, re, glob
sys.path.insert(0, os.path.dirname(__file__))
from tokhai_extract import extract_tokhai, is_tokhai_filename, _norm
from doc_classify import classify_file, pdf_text as _pdf_text
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

try:
    import pdfplumber
    PDF_READY = True
except ImportError:
    PDF_READY = False

DISK = '/Volumes/TOSHIBA EXT/HUATEX贷款材料'
L1 = os.path.join(DISK, 'L1-Raw Materials-Year2023.xlsx')
RM2023 = os.path.join(DISK, 'RM-Database/2023')
WS_COPY = '/Users/zhaoyuzhao/.openclaw/workspace-balance'
TS = datetime.now().strftime('%Y%m%d_%H%M')
TOL = 0.001

# 参数支持: 单月 "THÁNG 1-2023" | 范围 "1-6" | 单月数字 "3" | 空=全量
ARG = sys.argv[1] if len(sys.argv) > 1 else None
def resolve_months(arg):
    """返回 (目录名集合 or None表全量, tag标签)"""
    if not arg:
        return None, 'ALL'
    m = re.match(r'^(\d{1,2})\s*-\s*(\d{1,2})$', arg.strip())
    if m:  # 范围 1-6
        a, b = int(m.group(1)), int(m.group(2))
        return {f'THÁNG {i}-2023' for i in range(a, b + 1)}, f'{a}to{b}'
    if re.match(r'^\d{1,2}$', arg.strip()):  # 单月数字 3
        return {f'THÁNG {int(arg)}-2023'}, f'M{int(arg)}'
    return {arg}, re.sub(r'[^0-9A-Za-z]', '', arg)  # 完整目录名
MONTH_SET, tag = resolve_months(ARG)
OUT = os.path.join(DISK, f'Outcome1-v2.6-{tag}-{TS}.xlsx')
OUT_WS = os.path.join(WS_COPY, f'Outcome1-v2.6-{tag}-{TS}.xlsx')

# ================= HS Code → 用途分类（海关章节，权威依据）=================
# 纺织原料类 HS 章节: 50-60 (纱线/织物/纤维)
# 机器设备类 HS 章节: 84-85 (机械/电气设备)
def classify_by_hs(hs):
    if not hs or len(hs) < 2:
        return None
    ch = hs[:2]
    if ch in ('50','51','52','53','54','55','56','57','58','59','60','61','62','63'):
        return '原材料'
    if ch in ('84','85','90','73','82'):
        return '设备'
    return None

def cross_validate(inv_cls, inv_prod, hs_cls, hs):
    """Invoice品名为主 + HS码交叉验证(Daryl确认口径)
    互证一致→高；仅一源→中/高；两源打架→存疑低；都无→待确认低"""
    hs_tag = f'HS{hs}' if hs else '无HS码'
    prod = f'品名:{inv_prod}' if inv_prod else '品名未命中'
    # 两源都有结论
    if inv_cls and hs_cls:
        if inv_cls == hs_cls:
            return inv_cls, f'{prod} + {hs_tag}互证', '高'
        # 混合装运：Invoice读出混合，HS只给主品类 → 以Invoice(更细)为准
        if inv_cls == '混合(待拆分)':
            return inv_cls, f'{prod}(含多品类) | {hs_tag}仅主类{hs_cls}', '中'
        # 真打架 → 标存疑
        return f'⚠️存疑({inv_cls}/{hs_cls})', f'{prod}判{inv_cls} vs {hs_tag}判{hs_cls}', '低'
    # 仅Invoice品名
    if inv_cls:
        return inv_cls, f'{prod}(品名判定); {hs_tag}未归类', '中'
    # 仅HS码(海关权威)
    if hs_cls:
        return hs_cls, f'{hs_tag}(海关编码); {prod}', '高'
    # 都无
    return '⚠️待确认', f'{hs_tag}无法归类 + {prod}', '低'

# ================= Invoice 品名内容判定（辅助）=================
EQUIP_KW = ['MACHINE','MÁY','浆纱机','验布机','水洗机','染色机','经编机','织机','一体机',
            'WARPING','SIZING','WEAVING','KNITTING','DYEING','设备','机器','配件','SPARE PART',
            '钢筘','盘头','储油槽','锅炉','立库','喷淋塔','试验机','MOTOR','PUMP','BOILER']
RAW_KW = ['YARN','POLYESTER','FABRIC','NYLON','FIBER','FILAMENT','胚布','成品','纱线','助剂',
          '染料','毛坯','布','DTEX','DENIER','GREIGE','WOVEN','SPANDEX','COTTON']

def classify_by_invoice(dp, files):
    """打开 Invoice/销售合同 PDF 读品名 → 设备/原材料"""
    inv_files = []
    for f in files:
        fn = f.lower()
        if fn.endswith('.pdf') and any(k in f for k in
            ['合同','SC ','FULL SET','SHIPPING DOCS','清关','中英文明细','INVOICE','明细']):
            inv_files.append(f)
    text = ''
    for f in inv_files[:3]:
        text += _pdf_text(os.path.join(dp, f)) + '\n'
    tu = text.upper()
    e = [k for k in EQUIP_KW if k.upper() in tu]
    r = [k for k in RAW_KW if k.upper() in tu]
    # 提取品名行(Name of Commodity / description of goods 附近)
    prod = ''
    for line in text.split('\n'):
        lu = line.upper()
        if any(k in lu for k in ['POLYESTER','YARN','MACHINE','FABRIC','FILAMENT','NYLON']):
            prod = line.strip()[:40]
            break
    if e and r:
        return '混合(待拆分)', prod, '中'
    if e:
        return '设备', prod or ','.join(e), '高'
    if r:
        return '原材料', prod or ','.join(r), '高'
    return None, prod, ''

# ================= 文件内容分类用共享 classify_file (doc_classify) =================

# ================= 读 L1（唯一起点）=================
wb = openpyxl.load_workbook(L1, data_only=True)
ws = wb.active
l1_rows = []
for row in ws.iter_rows(min_row=4, values_only=True):
    if not row or len(row) < 6 or not row[2]:
        continue
    inv = str(row[2]).strip()
    if inv == 'None':
        continue
    try: amt = float(row[5]) if row[5] is not None else None
    except (ValueError, TypeError): amt = None
    d = row[4]
    if isinstance(d, datetime): dstr = d.strftime('%Y/%m/%d')
    else: dstr = str(d).replace('-','/').split(' ')[0] if d else ''
    l1_rows.append({'stt': row[0], 'supplier': str(row[1]).strip() if row[1] else '',
                    'inv': inv, 'l1_date': dstr, 'amount': amt})
wb.close()
print(f'L1 发票总数: {len(l1_rows)}')

# ================= 建 RM-Database 目录索引(发票号→目录路径) =================
# 遍历所有月份目录，建 发票号token → dirpath 映射，供 L1 发票查找
months = sorted(MONTH_SET) if MONTH_SET else sorted(os.listdir(RM2023))
dir_paths = []
for m in months:
    mp = os.path.join(RM2023, m)
    if not os.path.isdir(mp): continue
    for d in sorted(os.listdir(mp)):
        dp = os.path.join(mp, d)
        if os.path.isdir(dp):
            dir_paths.append((d, dp))

def find_dir_for_invoice(inv):
    """为 L1 发票号找 RM-Database 目录：目录名去序号前缀后，token 精确等于发票号"""
    matches = []
    for d, dp in dir_paths:
        core = re.sub(r'^\s*\d+\.\s*', '', d).strip()
        tokens = re.split(r'\s+', core)
        if inv in tokens:
            matches.append((d, dp))
    return matches

# ================= 主流程：只遍历 L1 =================
out_rows = []
scope_stts = set()   # 若指定月份，只输出该月份能在RM找到目录的发票
for rec in l1_rows:
    inv = rec['inv']
    matches = find_dir_for_invoice(inv)
    # 月份过滤：指定月份时，只保留在该月份目录中出现的发票
    if MONTH_SET and not matches:
        continue

    so = date = grade = note = ''
    cls = src = cf = ''
    complete = ''
    if not matches:
        # L1有但RM找不到目录
        so, date, grade = '', '', '❌未匹配'
        cls, src, cf = '⚠️待确认', 'RM-Database无对应目录', '低'
        complete = '⚠️目录缺失'
        note = 'RM-Database中无此发票目录'
    else:
        dp = matches[0][1]
        files = [f for f in os.listdir(dp) if os.path.isfile(os.path.join(dp, f))]
        # 分类文件 + 收集通关ToKhai
        cats = []
        cleared = []
        for f in files:
            c, r = classify_file(os.path.join(dp, f))
            cats.append(c)
            if c == 'TOKHAI_QDTQ' and r is not None and r['is_cleared']:
                cleared.append(r)
        # ToKhai 勾稽（方案B：金额完全相等）
        picked = [c for c in cleared if c['amount'] is not None
                  and rec['amount'] is not None and abs(c['amount']-rec['amount'])<TOL]
        if len(picked) == 1:
            so, date, grade = picked[0]['so_to_khai'], picked[0]['ngay_ymd'], '✅金额一致'
            hs = picked[0]['hs_code']
        elif len(picked) > 1:
            so = ';'.join(p['so_to_khai'] for p in picked)
            date = ';'.join(p['ngay_ymd'] for p in picked)
            grade = f'⚠️分批报关{len(picked)}单'
            note = '同发票多张通关单金额均符,已全列'
            hs = picked[0]['hs_code']
        else:
            so = date = ''
            grade = '❌未匹配'
            note = '⚠️无金额勾稽通关报关单-待人工复核'
            hs = cleared[0]['hs_code'] if cleared else ''
        # 三件套完整性
        has_inv = any(c in ('INVOICE_PDF','TOKHAI_QDTQ') for c in cats)
        has_bl = any(c in ('BL','SWB','AWB') for c in cats)
        has_tokhai = 'TOKHAI_QDTQ' in cats
        miss = []
        if not has_inv: miss.append('Invoice')
        if not has_bl: miss.append('B/L')
        if not has_tokhai: miss.append('ToKhai')
        complete = '✅完整' if (has_inv and has_bl) else '⚠️缺'+','.join(miss)
        # 用途分类: Invoice品名为主 + HS码交叉验证(Daryl确认口径)
        hs_cls = classify_by_hs(hs)
        inv_cls, inv_prod, _ = classify_by_invoice(dp, files)
        cls, src, cf = cross_validate(inv_cls, inv_prod, hs_cls, hs)

    out_rows.append({'stt': rec['stt'], 'inv': inv, 'so': so, 'date': date,
                     'amount': rec['amount'], 'l1_date': rec['l1_date'], 'supplier': rec['supplier'],
                     'cls': cls, 'src': src, 'cf': cf, 'complete': complete,
                     'grade': grade, 'note': note})

# 排序: STT 升序
def stt_key(r):
    try: return (0, int(r['stt']))
    except (ValueError, TypeError): return (1, 0)
out_rows.sort(key=stt_key)

# ================= 写 Outcome1 =================
wbo = openpyxl.Workbook(); sh = wbo.active
sh.title = '银行放款明细表'
headers = ['序号','发票号','关单号','报关日期','金额(USD)','日期','供应商',
           '贷款用途分类','分类依据','置信度','三件套完整性','ToKhai勾稽','备注']
hfill = PatternFill('solid', fgColor='4472C4')
hfont = Font(color='FFFFFF', bold=True, size=10)
thin = Side(style='thin', color='BFBFBF'); border = Border(thin,thin,thin,thin)
for c,h in enumerate(headers,1):
    cell=sh.cell(1,c,h); cell.fill=hfill; cell.font=hfont
    cell.alignment=Alignment(horizontal='center',vertical='center'); cell.border=border
r=2
for row in out_rows:
    vals=[row['stt'],row['inv'],row['so'],row['date'],row['amount'],row['l1_date'],
          row['supplier'],row['cls'],row['src'],row['cf'],row['complete'],row['grade'],row['note']]
    for c,v in enumerate(vals,1):
        cell=sh.cell(r,c,v); cell.border=border
        cell.alignment=Alignment(horizontal='center' if c in(1,3,4,10) else 'left',vertical='center')
        if c==5 and isinstance(v,(int,float)): cell.number_format='#,##0.00'
    r+=1
widths=[7,17,16,13,13,12,24,14,26,8,18,15,26]
for c,w in enumerate(widths,1):
    sh.column_dimensions[openpyxl.utils.get_column_letter(c)].width=w
sh.freeze_panes='A2'
wbo.save(OUT)
import shutil
shutil.copy2(OUT, OUT_WS)

# 统计
lo = sum(1 for r in out_rows if r['cf']=='低')
hi = sum(1 for r in out_rows if r['cf']=='高')
matched = sum(1 for r in out_rows if r['grade'].startswith('✅') or '分批' in r['grade'])
print(f'✅ 生成: {OUT}')
print(f'✅ workspace副本: {OUT_WS}')
print(f'输出行数(严格=L1范围): {len(out_rows)}')
print(f'ToKhai勾稽成功: {matched} | 高置信分类: {hi} | 低置信: {lo}')
