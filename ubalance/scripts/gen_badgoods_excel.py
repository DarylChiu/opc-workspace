#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""生成《越南账外不良品处理方案》Excel 版"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---- 样式 ----
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="微软雅黑", bold=True, size=14, color="1F4E79")
SUB_FONT = Font(name="微软雅黑", bold=True, size=11, color="1F4E79")
BODY_FONT = Font(name="微软雅黑", size=10)
RED_FONT = Font(name="微软雅黑", size=10, color="C00000")
ORANGE_FONT = Font(name="微软雅黑", size=10, color="BF6000")
GREEN_FONT = Font(name="微软雅黑", size=10, color="2E7D32")
WRAP = Alignment(wrap_text=True, vertical="top")
WRAP_C = Alignment(wrap_text=True, vertical="center", horizontal="center")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
ZEBRA = PatternFill("solid", fgColor="F2F7FB")

def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = WRAP_C
        cell.border = BORDER

def style_body(ws, row, ncols, fonts=None):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = fonts[c - 1] if fonts else BODY_FONT
        cell.alignment = WRAP
        cell.border = BORDER
        if row % 2 == 0:
            cell.fill = ZEBRA

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def add_title(ws, text, span):
    ws.cell(row=1, column=1, value=text).font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)

wb = openpyxl.Workbook()

# ============ Sheet 1: 执行摘要 ============
ws = wb.active
ws.title = "执行摘要"
add_title(ws, "越南账外不良品（约200吨）处理方案 · 执行摘要", 3)
ws.cell(row=2, column=1, value="编制: Balance（算点小账） | 日期: 2026-08-17 | 源起: 2026-08-15 初步方案 + 方案4/5/6深挖").font = Font(name="微软雅黑", size=9, italic=True, color="808080")
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)

rows = [
    ("标的", "约 200 吨「账外不良品」，已按 BOM 损耗率出表（账面不存在、实物在库）。"),
    ("总开关", "先解决「账外」——来源合法化（补账完税 / 正式报废）是所有报关出口方案的前置条件，不解决则一切方案都带雷。"),
    ("货物出路", "三条：A 越南内销（🟢最稳）、B 出口香港（🟡资金出境最干净）、C 直出大陆（🔴固废红线）。"),
    ("资金/利润流转", "方案5（服务费通道，主力提取）+ 方案6（保理/福费廷，资金回笼加速器）为 Daryl 8/15 定调组合；方案4（预付+退款）为领导方向；股息分红（0%预提税，先交20%CIT）最干净合规备选。"),
    ("我的判断", "大概率 A+B 组合打底，货物端三角转口让毛利沉淀香港/新加坡；直出大陆仅在「非固废属性 + Form E 原产地」双条件成立时才考虑。"),
]
r = 4
ws.cell(row=r, column=1, value="项目").font = SUB_FONT
ws.cell(row=r, column=2, value="内容").font = SUB_FONT
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
style_header(ws, r, 3)
r += 1
for k, v in rows:
    ws.cell(row=r, column=1, value=k)
    ws.cell(row=r, column=2, value=v)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    style_body(ws, r, 3)
    ws.cell(row=r, column=1).font = Font(name="微软雅黑", size=10, bold=True)
    r += 1

ws.cell(row=r+1, column=1, value="合规声明: 本文件是决策辅助研究，逐条标注法律风险与代价。任何落地路径均须经越南本地律师/审计师复核；涉及报关、税务、外汇的结论以主管机关现行文件为准。").font = Font(name="微软雅黑", size=9, italic=True, color="C00000")
ws.merge_cells(start_row=r+1, start_column=1, end_row=r+1, end_column=3)
set_widths(ws, [18, 70, 15])

# ============ Sheet 2: 问题定义 ============
ws = wb.create_sheet("问题定义")
add_title(ws, "一、问题定义（口径钉死）", 3)
hdr = ["项", "状态", "备注"]
for i, h in enumerate(hdr, 1):
    ws.cell(row=3, column=i, value=h)
style_header(ws, 3, 3)
data = [
    ("标的", "约200吨账外「不良品」", "待确认：棉/化纤/混纺？布匹/纱线/边角料/成衣？"),
    ("账务状态", "已按BOM损耗率出表（损耗核销）", "账面不存在、实物在库"),
    ("监管状态", "自述已脱离保税/海关/税务控制", "待确认：是否加工贸易料件、核销文件"),
    ("目标", "报关销售回中国大陆或香港", "评估 ①可行性 ②税负成本 ③资金流转"),
]
r = 4
for a, b, c in data:
    ws.cell(row=r, column=1, value=a)
    ws.cell(row=r, column=2, value=b)
    ws.cell(row=r, column=3, value=c)
    style_body(ws, r, 3)
    ws.cell(row=r, column=1).font = Font(name="微软雅黑", size=10, bold=True)
    r += 1
ws.cell(row=r+1, column=1, value="⚠️ 第一个要钉死的事实：这批货是加工贸易（保税料件）产生的，还是内购料/内销生产产生的？法律框架完全不同。").font = RED_FONT
ws.merge_cells(start_row=r+1, start_column=1, end_row=r+1, end_column=3)
set_widths(ws, [16, 40, 45])

# ============ Sheet 3: 法规地图 ============
ws = wb.create_sheet("法规地图")
add_title(ws, "二、关键法规地图", 3)
hdr = ["区域", "条目", "要点"]
for i, h in enumerate(hdr, 1):
    ws.cell(row=3, column=i, value=h)
style_header(ws, 3, 3)
data = [
    ("越南·出口", "出口 VAT 0%", "出口商品适用0%税率（Luật GTGT Điều 8），前提=真实出口报关+出口单据+按期收汇"),
    ("越南·出口", "加工贸易废料/次级品处理", "TT 38/2015/TT-BTC Điều 71 + TT 39/2018 修订 + NĐ 69/2018 Điều 44 + CV 2687/TCHQ-TXNK 2021；合法出路仅三条：①内销 ②再出口 ③销毁；内销免进口税条件=废料量≤进口料件总量3%（超出补进口税）"),
    ("越南·出口", "发票/汇率", "按 TT 99/2025/TT-BTC 汇率折算；电子发票按 NĐ 70/2025/NĐ-CP"),
    ("越南·出口", "外汇收汇", "出口货款须限期回流越南并核销（逾期=罚款/强制结汇，具体天数按NHNN现行规定待核实）"),
    ("越南·出口", "报关前提", "货物有合法来源：发票、账务、实物三流一致——这是账外货的死穴"),
    ("中国·进口", "固体废物禁令（最硬红线）", "2021-01-01起全面禁止固体废物进口（生态环境部公告2020年第4号 + 新《固废法》）；纺织品废料 HS 5202/5505/5301 = 固废 → 禁止进口，申报即退运+处罚"),
    ("中国·进口", "次级品可入", "次级品/残次品（仍有使用价值，按织物品名申报如5208-5212）→ 可正常进口，交关税+VAT；品名与实物状态必须扛得住查验"),
    ("中国·进口", "关税/原产地", "中国-东盟 Form E → 纺织品协定税率通常0%；RCEP同理；无原产地证→MFN关税（棉织物约8%档）；增值税13%（进口环节买家缴，可抵扣）"),
    ("中国·进口", "原产地疑点", "若原料是保税料件，越南加工后次级品是否满足Form E原产地规则——待确认，不满足则只能走MFN"),
    ("香港·进口", "税负最轻", "香港无关税、无增值税；对固体废物进口也有管制（环保署许可），次级品/可再用商品不受限；香港为中间贸易/离岸结算平台典型地位"),
]
r = 4
for a, b, c in data:
    ws.cell(row=r, column=1, value=a)
    ws.cell(row=r, column=2, value=b)
    ws.cell(row=r, column=3, value=c)
    style_body(ws, r, 3)
    ws.cell(row=r, column=1).font = Font(name="微软雅黑", size=10, bold=True)
    ws.cell(row=r, column=2).font = Font(name="微软雅黑", size=10, bold=True)
    r += 1
set_widths(ws, [14, 30, 60])

# ============ Sheet 4: 来源合法化 ============
ws = wb.create_sheet("来源合法化")
add_title(ws, "三、总开关：来源合法化（所有方案的前置问题）", 3)
hdr = ["路径", "操作", "成本/风险"]
for i, h in enumerate(hdr, 1):
    ws.cell(row=3, column=i, value=h)
style_header(ws, 3, 3)
data = [
    ("A. 补账完税（合规）", "废料/次级品重新入账→确认收入→交CIT 20% + 内销增值税或出口0%", "税负=收入全额×20% CIT（成本已出表，无成本抵扣）；但从此干净", GREEN_FONT),
    ("B. 第三方收购洗白（灰色）", "找越南废品商「收购」→取得发票→再出口", "发票成本（票面价×税点）+ 虚开发票风险（越南近年严打，税局发票对碰）", ORANGE_FONT),
    ("C. 全程账外非正规（高危）", "不报关/低报/地下渠道", "=走私+逃税+固废违规，越南刑事风险（逃税超1亿VND可入刑），不建议", RED_FONT),
]
r = 4
for a, b, c, f in data:
    ws.cell(row=r, column=1, value=a)
    ws.cell(row=r, column=2, value=b)
    ws.cell(row=r, column=3, value=c)
    style_body(ws, r, 3)
    ws.cell(row=r, column=1).font = Font(name="微软雅黑", size=10, bold=True)
    ws.cell(row=r, column=3).font = f
    r += 1
ws.cell(row=r+1, column=1, value="提醒: 越南税务稽查近年高发项=废料/副产物收入不入账（账外卖废料被查=补税+罚款0.5-2倍+滞纳金，达金额入刑）。200吨量级走C路径，风险极高。").font = RED_FONT
ws.merge_cells(start_row=r+1, start_column=1, end_row=r+1, end_column=3)
set_widths(ws, [24, 45, 40])

# ============ Sheet 5: 货物方案 A/B/C ============
ws = wb.create_sheet("货物方案ABC")
add_title(ws, "四、货物出口方案 A/B/C 对比", 3)

# 对比总表
hdr = ["维度", "A 内销/经纪", "B 香港转口", "C 直出大陆"]
for i, h in enumerate(hdr, 1):
    ws.cell(row=3, column=i, value=h)
style_header(ws, 3, 4)
compare = [
    ("可行性", "🟢 最高", "🟡 中", "🟡 中低（固废红线）"),
    ("越南税负", "VAT10%+CIT20%（入账）", "VAT 0%", "VAT 0%"),
    ("进口端税负", "—", "0", "0%（FormE）/8%+13%VAT"),
    ("资金目的地", "留在越南", "境外留存（最灵活）", "回流越南/人民币"),
    ("综合风险", "🟡 中", "🟡 中", "🔴 高"),
    ("单吨净价预判", "低（越南本地收购价）", "中（HK价）", "高（大陆价，若FormE）"),
]
r = 4
for row in compare:
    for c, v in enumerate(row, 1):
        ws.cell(row=r, column=c, value=v)
    style_body(ws, r, 4)
    ws.cell(row=r, column=1).font = Font(name="微软雅黑", size=10, bold=True)
    r += 1

# 各方案细节
r += 2
ws.cell(row=r, column=1, value="各方案路径细节").font = SUB_FONT
r += 1
detail_hdr = ["方案", "路径", "可行性", "税负", "资金", "适合", "风险"]
for i, h in enumerate(detail_hdr, 1):
    ws.cell(row=r, column=i, value=h)
style_header(ws, r, 7)
r += 1
details = [
    ("A 越南境内变现", "直接内销给越南废料/次级品收购商（拿越南盾），或出口给越南经纪商由其转口", "🟢最高", "合规=VAT10%+CIT20%；灰色=0税但全风险", "越南盾收款，资金留在越南", "只要变现、不在乎资金在越；或越南公司有人民币/美金需求可对冲", "🟡中"),
    ("B 出口香港", "越南出口报关（品名=次级品织物，非废料）→香港公司收货", "🟡中", "越南出口VAT0%；香港0%", "HK收USD/HKD→留存境外资金池", "目标=资金出境+境外留存；或HK已有贸易主体", "🟡中"),
    ("C 直出大陆", "越南出口（次级品）→中国进口报关（Form E或MFN）→国内销售", "🟡中低", "越南VAT0%；中国关税0%(FormE)/8%(MFN)+VAT13%", "人民币跨境或USD TT→回流越南或经HK中转", "最终市场在中国、买家现成、愿走正规报关", "🔴高"),
]
for row in details:
    for c, v in enumerate(row, 1):
        ws.cell(row=r, column=c, value=v)
    style_body(ws, r, 7)
    ws.cell(row=r, column=1).font = Font(name="微软雅黑", size=10, bold=True)
    r += 1
set_widths(ws, [22, 45, 18, 32, 30, 40, 16])

# ============ Sheet 6: 资金通道 4/5/6 ============
ws = wb.create_sheet("资金通道456")
add_title(ws, "五、资金与利润流转通道（方案4/5/6）", 3)
hdr = ["维度", "4 预付+退款/诉讼", "5 服务费", "6 保理"]
for i, h in enumerate(hdr, 1):
    ws.cell(row=3, column=i, value=h)
style_header(ws, 3, 4)
compare = [
    ("功能", "提取", "提取", "利润截留+金融化/资金回笼"),
    ("成本", "≈0%", "5-10%", "8-12%（折扣）"),
    ("程序", "诉讼/仲裁（重）", "备案+付汇（轻）", "保理协议（轻）"),
    ("实质风险", "关联诉讼真实性", "服务实质审查", "关联保理折扣合理性"),
    ("提取量", "大", "大", "小（折扣）"),
    ("定位", "领导方向", "主力提取", "辅助/资金回笼"),
]
r = 4
for row in compare:
    for c, v in enumerate(row, 1):
        ws.cell(row=r, column=c, value=v)
    style_body(ws, r, 4)
    ws.cell(row=r, column=1).font = Font(name="微软雅黑", size=10, bold=True)
    r += 1

r += 2
ws.cell(row=r, column=1, value="各通道链条说明").font = SUB_FONT
r += 1
chain_hdr = ["方案", "链条", "关键前提/死穴"]
for i, h in enumerate(chain_hdr, 1):
    ws.cell(row=r, column=i, value=h)
style_header(ws, r, 3)
r += 1
chains = [
    ("方案4 预付+退款", "HK预付50%→发部分货→质量不合格终止→退款+违约金。预付款退款是常规商业操作、银行审核轻、不需裁决书；超预付部分靠合同违约金条款支撑。可与诉讼赔付叠加两段式：预付款走小额、诉讼走大额。", "程序重（需诉讼/仲裁支撑）"),
    ("方案5 服务费", "货物出口→货款回流越南→越南实体向HK/新加坡关联方付服务费→FCT代扣代缴→付汇出境。", "服务必须真实（境外方有人员/报告/交付物）、费率市场价、合同+备案+付汇三流一致；空壳收服务费=重定性+补税（TT 96/2015）"),
    ("方案6 保理/福费廷", "壳公司赊账出口(60-90天)→应收款卖断给境外保理商→保理商折价付现（出口收汇完成）→HK买家到期付款给保理商（境外闭环）。折扣8-12%留境外；若保理商受控=折扣率即利润转移载体。", "死穴：银行先掏钱买应收款→需授信审批（信用/流水/报表）；壳公司四项全无→只能走受控保理商（本质=关联保理/内部资金池，性质已变）"),
]
for a, b, c in chains:
    ws.cell(row=r, column=1, value=a)
    ws.cell(row=r, column=2, value=b)
    ws.cell(row=r, column=3, value=c)
    style_body(ws, r, 3)
    ws.cell(row=r, column=1).font = Font(name="微软雅黑", size=10, bold=True)
    r += 1

# 方案5税负明细表
r += 1
ws.cell(row=r, column=1, value="方案5 税负明细（TT 103/2014 FCT + 越南-HK DTA）").font = SUB_FONT
r += 1
tax_hdr = ["服务类型", "CIT预提", "VAT", "净成本"]
for i, h in enumerate(tax_hdr, 1):
    ws.cell(row=r, column=i, value=h)
style_header(ws, r, 4)
r += 1
tax_rows = [
    ("管理/咨询/市场服务", "5%", "5%（可抵扣）", "≈5%"),
    ("技术转让/特许权使用费", "10%（DTA→7.5%）", "5%（可抵扣）", "≈7.5-10%"),
]
for row in tax_rows:
    for c, v in enumerate(row, 1):
        ws.cell(row=r, column=c, value=v)
    style_body(ws, r, 4)
    r += 1
set_widths(ws, [24, 45, 55])

# ============ Sheet 7: 组合建议 ============
ws = wb.create_sheet("组合建议")
add_title(ws, "六、组合建议（2026-08-15 Daryl 定调）", 3)
data = [
    ("1. 货物端", "方案 A+B 组合打底（越南侧小额内销 + 大头出口香港），三角转口让毛利沉淀香港/新加坡，减少越南账面利润（少交20% CIT）。"),
    ("2. 资金端", "方案5（服务费，管理咨询+技术服务双名目分散）做提取主力 + 方案6（保理）做资金回笼辅助 + 方案4 留给领导层决策。"),
    ("3. 合规备选（最干净）", "股息分红汇出=0%预提税（KPMG/越南简报多源确认）——但壳公司利润需先交20% CIT + 审计。「贵但最干净」，作为合规兜底。"),
    ("4. 唯一真问题", "货最终进不进中国大陆？进大陆=次级品报关+扛查验（固废红线还在），多脚只能解决资金追踪、解决不了品名；不进大陆（留HK/东南亚消化）=红线整个消失，结构能更简单。"),
]
r = 3
for k, v in data:
    ws.cell(row=r, column=1, value=k)
    ws.cell(row=r, column=2, value=v)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    style_body(ws, r, 3)
    ws.cell(row=r, column=1).font = Font(name="微软雅黑", size=10, bold=True)
    r += 1
set_widths(ws, [24, 65, 15])

# ============ Sheet 8: 风险清单 ============
ws = wb.create_sheet("风险清单")
add_title(ws, "七、风险清单（按爆炸当量排序）", 3)
hdr = ["等级", "风险", "说明"]
for i, h in enumerate(hdr, 1):
    ws.cell(row=3, column=i, value=h)
style_header(ws, 3, 3)
risks = [
    ("🔴 高", "中国固废禁令（方案C独有）", "纺织品废料=禁止进口固废；以次级品申报=品名造假风险，被查=退运+罚款+企业失信，金额大=走私罪", RED_FONT),
    ("🔴 高", "越南逃税刑责", "账外收入不申报，逃税额>1亿VND可刑事追诉；越南税局废料稽查专项", RED_FONT),
    ("🟠 中高", "海关核销对账", "已按BOM损耗核销的料件出现实物出口，若报关品名/数量与核销记录矛盾→海关追溯核销真实性（加工贸易稽查）", ORANGE_FONT),
    ("🟠 中高", "虚开发票（来源洗白路径B）", "越南2026年电子发票全链条对碰，买票洗白=发票犯罪", ORANGE_FONT),
    ("🟠 中高", "外汇违规", "出口收汇不按期回流=NHNN罚款+强制结汇；地下钱庄出境=洗钱风险", ORANGE_FONT),
    ("🟡 中", "Form E 原产地不符", "保税料件加工品可能不满足原产地规则→关税从0%变8%+补税", BODY_FONT),
]
r = 4
for a, b, c, f in risks:
    ws.cell(row=r, column=1, value=a)
    ws.cell(row=r, column=2, value=b)
    ws.cell(row=r, column=3, value=c)
    style_body(ws, r, 3)
    ws.cell(row=r, column=1).font = f
    ws.cell(row=r, column=2).font = Font(name="微软雅黑", size=10, bold=True)
    r += 1
set_widths(ws, [14, 30, 60])

# ============ Sheet 9: 待确认信息 + 下一步 ============
ws = wb.create_sheet("待确认与下一步")
add_title(ws, "八、待确认信息 & 九、下一步动作", 2)
ws.cell(row=2, column=1, value="待确认信息（决定方案深度）").font = SUB_FONT
confirm = [
    "货物形态与材质：棉/化纤/混纺？布匹/纱线/边角料/成衣？规格单？→ 定 HS 归类",
    "料件属性：加工贸易（保税）料件还是内购料？核销文件/损耗率审批单有没有？",
    "价值量级：200吨估算货值（人民币）？中国/香港是否有现成买家？",
    "资金最终目的地：回中国境内？留境外？留越南？→ 决定 A/B/C 怎么选",
    "公司风险偏好：愿意补账完税（交CIT），还是必须完全账外？→ 决定走合规版还是灰色版",
]
r = 3
for i, v in enumerate(confirm, 1):
    ws.cell(row=r, column=1, value=f"{i}. {v}")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    style_body(ws, r, 2)
    r += 1

r += 1
ws.cell(row=r, column=1, value="下一步动作").font = SUB_FONT
r += 1
next_steps = [
    "Daryl 确认「货最终进不进中国大陆」——定结构简繁",
    "补齐第八节资料（货物形态/料件属性/货值/资金目的地/风险偏好）",
    "补齐后做 HS 级细算（税负精算到单吨人民币净回款）+ 方案选择树",
    "各通道出全链路金额测算 + 落地文书清单（FCT备案/完税证明/合同）",
    "落地前经越南本地律师/审计师复核",
]
for i, v in enumerate(next_steps, 1):
    ws.cell(row=r, column=1, value=f"☐ {v}")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    style_body(ws, r, 2)
    r += 1
set_widths(ws, [100, 15])

# 冻结首行 + 保存
for ws in wb.worksheets:
    ws.sheet_view.showGridLines = False

out = "/Users/zhaoyuzhao/.openclaw/workspace-balance/不良品处理方案-完整版-20260817.xlsx"
wb.save(out)
print("SAVED:", out)
